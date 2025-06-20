import os
import time
import json
from difflib import SequenceMatcher
from typing import List, Dict, Any, Optional

import pandas as pd
import requests
from openpyxl import load_workbook
from tqdm import tqdm

###############################################################################
# 工具函数
###############################################################################

def get_access_token() -> Optional[str]:
    """ 
    调用 OAuth2 *Client Credentials* 模式获取 **access_token**，有效期 1 小时。

    Returns
    -------
    str | None
        成功时返回 token 字符串；失败时打印错误并返回 ``None``。
    """
    url = "https://www.ymgal.games/oauth/token"
    data = {
        "grant_type": "client_credentials",
        "client_id": "ymgal",  # 固定 client_id，由月幕平台提供
        "client_secret": "luna0327",  # 固定 client_secret，由月幕平台提供
        "scope": "public"  # 只申请公开数据权限
    }
    response = requests.post(url, data=data)

    if response.status_code == 200:
        return response.json().get("access_token")

    # 失败时输出详细信息，方便排查
    print("获取 token 失败:", response.status_code, response.text)
    return None

# ---------------------------------------------------------------------------
# 搜索相关辅助
# ---------------------------------------------------------------------------

def parse_search_response(response: requests.Response) -> List[Dict[str, Any]]:
    """ 
    解析 *search-game* 接口返回，提取游戏及其会社信息。 

    参数
    ----
    response : requests.Response
        月幕 *search-game* API 响应对象。

    Returns
    -------
    list[dict]
        解析后的结果列表，每个元素均包含：
        - ``name``：日文 / 英文原名
        - ``chineseName``：中文名(可能为空)
        - ``ym_id``：月幕游戏 ID
        - ``score``：月幕算法打分 (匹配度)
        - ``orgId`` / ``orgName`` / ``orgWebsite`` / ``orgDescription``：会社信息
    """
    try:
        response_data = response.json()
        # --- 调试输出，可根据需要关闭 ------------------------------------
        print("\n完整 API 响应：")
        print(json.dumps(response_data, indent=2, ensure_ascii=False))
        # ------------------------------------------------------------------
        results = response_data.get("data", {}).get("result", [])
    except Exception as exc:
        print("解析 response 失败：", exc)
        return []

    parsed: List[Dict[str, Any]] = []
    for item in results:
        # 1️⃣ 解析匹配分数，默认 0.0
        try:
            score = float(item.get("score", 0))
        except (ValueError, TypeError):
            score = 0.0

        # 2️⃣ 解析会社信息，API 有时嵌套在 ``org``，有时散落在顶层
        org_info = item.get("org", {}) or {
            "id": item.get("orgId", ""),
            "name": item.get("orgName", ""),
            "website": item.get("orgWebsite", ""),
            "description": item.get("orgDescription", "")
        }

        if org_info:
            print(f"找到会社信息：{org_info.get('name', '')}")

        parsed.append({
            "name": item.get("name", ""),
            "chineseName": item.get("chineseName", ""),
            "ym_id": item.get("id", ""),
            "score": round(score, 4),
            "orgId": org_info.get("id", ""),
            "orgName": org_info.get("name", ""),
            "orgWebsite": org_info.get("website", ""),
            "orgDescription": org_info.get("description", "")
        })

    return parsed

def search_ym_top_matches(
    keyword: str,
    token_ref: Dict[str, str],
    top_k: int = 3,
    threshold: float = 0.8
) -> List[Dict[str, Any]]:
    """ 
    根据 *keyword* 在月幕搜索游戏并返回最相关的前 ``top_k`` 条结果。

    特性：
    --------
    - **Token 自动刷新**：若接口返回 401 则重新获取一次 token，最多重试 4 次。
    - **阈值过滤**：若最高得分 >= ``threshold`` 则只返回 1 条最优匹配。

    参数
    ----
    keyword : str
        待搜索的 Bangumi 游戏名称。
    token_ref : dict
        形如 ``{"value": <token>}`` 的可变字典，用于在内部更新失效 token。
    top_k : int, default=3
        未触发阈值过滤时，返回结果数。
    threshold : float, default=0.8
        最高得分超过该阈值时，视为高度一致，仅返回首条。

    Returns
    -------
    list[dict]
        解析后的匹配结果列表 (可能为空)。
    """

    def _make_request(token: str) -> requests.Response:
        """内部封装：携带 token 调用 search-game 接口。"""
        url = "https://www.ymgal.games/open/archive/search-game"
        params = {
            "mode": "list",
            "keyword": keyword,
            "pageNum": 1,
            "pageSize": 20,
            "includeOrg": "true"
        }
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "version": "1"
        }
        return requests.get(url, params=params, headers=headers, timeout=10)

    # --- 主流程：最多尝试 4 次 --------------------------------------------
    for attempt in range(4):
        token = token_ref["value"]
        response = _make_request(token)

        # 1. 请求成功 -> 解析
        if response.status_code == 200:
            matches = parse_search_response(response)
            matches = sorted(matches, key=lambda x: x["score"], reverse=True)

            # 阈值过滤逻辑
            if matches and matches[0]["score"] >= threshold:
                return matches[:1]
            return matches[:top_k]

        # 2. Token 失效 -> 刷新后重试
        elif response.status_code == 401:
            print("Token 失效，正在重新获取…")
            new_token = get_access_token()
            if new_token:
                token_ref["value"] = new_token
                continue
            print("重新获取 token 失败")
            return []

        # 3. 其它错误 -> 直接返回空
        else:
            print(f"搜索失败: {response.status_code}, {response.text}")
            return []

    # 超出重试次数
    return []

###############################################################################
# Excel 处理函数
###############################################################################

# 🌟 本节函数通过 ``pandas`` 与 ``openpyxl`` 实现数据的增量写入、
#    防覆盖写入以及临时文件兜底等高级需求。

EXCEL_COLUMNS_MATCHED = [
    "bgm_id", "bgm游戏", "日文名 (原始)", "中文名 (原始)",
    "name", "chineseName", "ym_id", "score",
    "orgId", "orgName", "orgWebsite", "orgDescription",
    "匹配来源"
]

EXCEL_COLUMNS_ORG = [
    "org_id", "name", "chineseName", "website", "description", "birthday"
]

def init_excel(output_file: str) -> None:
    """ 
    确保匹配结果文件存在；若不存在或损坏则创建带表头的新文件。
    """
    need_create = False
    if not os.path.exists(output_file):
        need_create = True
    else:
        try:
            load_workbook(output_file)
        except Exception:
            need_create = True

    if need_create:
        pd.DataFrame(columns=EXCEL_COLUMNS_MATCHED).to_excel(output_file, index=False)
        print(f"已初始化输出文件：{output_file}")

def init_org_excel(output_file: str) -> None:
    """类似 ``init_excel``，但针对会社信息文件。"""
    if not os.path.exists(output_file):
        pd.DataFrame(columns=EXCEL_COLUMNS_ORG).to_excel(output_file, index=False)
        print(f"已初始化会社信息文件：{output_file}")

def append_to_excel(row_data: List[Dict[str, Any]], output_file: str) -> None:
    """ 
    将 ``row_data`` 追加写入到 ``output_file``，支持自动创建及占用兜底。
    """
    try:
        df_new = pd.DataFrame(row_data)

        # 1️⃣ 文件不存在：直接写
        if not os.path.exists(output_file):
            df_new.to_excel(output_file, index=False)
            return

        # 2️⃣ 文件存在：读取 + 合并 + 写回
        try:
            df_exist = pd.read_excel(output_file)
            df_combined = pd.concat([df_exist, df_new], ignore_index=True)
            df_combined.to_excel(output_file, index=False)
        except PermissionError:  # 常见于文件被 Excel 占用
            temp_file = f"{output_file}.temp"
            df_new.to_excel(temp_file, index=False)
            print(f"原文件被占用，数据已保存到临时文件：{temp_file}")
    except Exception as exc:
        # 兜底打印 & 备份
        print(f"保存数据时发生错误: {exc}")
        backup_file = f"{output_file}.backup"
        pd.DataFrame(row_data).to_excel(backup_file, index=False)
        print(f"数据已保存到备用文件：{backup_file}")

def append_unmatched_to_excel(name: str, unmatched_file: str) -> None:
    """记录未匹配成功的 Bangumi 名称。"""
    df = pd.DataFrame([[name]], columns=["原始的未匹配bgm游戏名称"])
    if not os.path.exists(unmatched_file):
        df.to_excel(unmatched_file, index=False)
    else:
        with pd.ExcelWriter(unmatched_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            sheet = writer.book["Sheet1"]
            df.to_excel(writer, index=False, header=False, startrow=sheet.max_row)

def append_org_to_excel(org_info: Dict[str, Any], output_file: str) -> None:
    """将会社信息写入文件，逻辑同 ``append_to_excel``。"""
    append_to_excel([org_info], output_file)

###############################################################################
# 会社详细信息查询
###############################################################################

def get_organization_details(org_id: str, token: str) -> Optional[Dict[str, Any]]:
    """ 
    根据 ``org_id`` 向月幕查询会社详细资料。

    返回的字段包括：名称、中文名、官网、简介、成立日期等。
    若调用失败或字段缺失，则返回 ``None``。
    """
    url = "https://www.ymgal.games/open/archive"
    params = {"orgId": org_id}
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "version": "1"
    }

    try:
        # ---------- 调试信息 --------------
        print("\n正在获取会社信息… ID:", org_id)
        # ---------------------------------
        response = requests.get(url, params=params, headers=headers, timeout=10)

        if response.status_code == 200:
            data = response.json()
            org_data = data.get("data", {}).get("org", {})
            if not org_data:
                print("API 响应中未找到会社信息")
                return None

            # 按优先级提取官网地址，fallback 使用第一个
            website = ""
            if isinstance(org_data.get("website"), list):
                priority = ["homepage", "官网", "官方网站", "official website"]
                for title in priority:
                    for site in org_data["website"]:
                        if site.get("title", "").lower() == title.lower():
                            website = site.get("link", "")
                            break
                    if website:
                        break
                if not website and org_data["website"]:
                    website = org_data["website"][0].get("link", "")

            # 组装结果
            result = {
                "id": org_id,
                "name": org_data.get("name", ""),
                "chineseName": org_data.get("chineseName", ""),
                "website": website,
                "description": org_data.get("introduction", ""),
                "birthday": org_data.get("birthday", "")
            }
            return result

        if response.status_code == 401:  # token 失效, 交由外层处理
            print("公司信息获取时 token 失效")
            return None

        print(f"获取会社信息失败: {response.status_code}")
        return None

    except Exception as exc:
        print(f"获取会社信息时发生错误: {exc}")
        return None

###############################################################################
# 主流程：Bangumi -> 月幕 首次匹配
###############################################################################

def match_bgm_games_and_save(
    input_file: str = r"E:\学习资料\项目文件\BaiduTiebaSpider-main\ymproject2\bgm_archive_20250525 (1).xlsx",
    output_file: str = "ymgames_matched.xlsx",
    unmatched_file: str = "ymgames_unmatched.xlsx",
    org_output_file: str = "organizations_info.xlsx"
)-> None:
    """ 
    读取 Bangumi Excel -> 月幕搜索匹配 -> 写结果
    支持 **断点续跑** ：已处理过的 Bangumi 名称会跳过。
    """
    # 1. 读取 Bangumi 源文件
    df_bgm = pd.read_excel(input_file, engine="openpyxl")
    print(f"DEBUG: 识别到的 Excel 列名：{df_bgm.columns.tolist()}")
    
    if "日文名" not in df_bgm.columns or "中文名" not in df_bgm.columns:
        raise ValueError("Excel 中必须包含 '日文名' 和 '中文名' 列")

    game_names_cn: List[str] = df_bgm["中文名"].dropna().astype(str).tolist()
    
    # 2. 加载已处理过的 ID (用于断点续跑)
    processed_ids: set[Any] = set()
    if os.path.exists(output_file):
        try:
            df_exist = pd.read_excel(output_file, engine="openpyxl")
            if 'bgm_id' in df_exist.columns:
                processed_ids = set(df_exist["bgm_id"].dropna().astype(str))
            else:
                print("警告: 输出文件中未找到 'bgm_id' 列，断点续跑可能不准确。")
        except Exception as exc:
            print("读取已匹配文件失败，将重新创建：", exc)

    # 3. 初始化输出文件 & token
    token_ref = {"value": get_access_token()}
    if not token_ref["value"]:
        print("无法获取 token，流程终止")
        return

    init_excel(output_file)
    init_org_excel(org_output_file)

    # 4. 加载已有会社信息到内存，避免重复查询
    processed_orgs: Dict[str, Dict[str, Any]] = {}
    if os.path.exists(org_output_file):
        try:
            org_df = pd.read_excel(org_output_file, engine="openpyxl")
            for _, row in org_df.iterrows():
                org_id = str(row["org_id"])
                if pd.notna(org_id):
                    processed_orgs[org_id] = {
                        "info": row.to_dict(),
                        "retry_count": 0
                    }
        except Exception as exc:
            print("读取会社信息文件失败，将重新创建：", exc)

    # 5. 遍历 Bangumi 行并匹配
    for idx, row in tqdm(df_bgm.iterrows(), total=len(df_bgm), desc="处理游戏"):
        bgm_id = str(row['id']) if 'id' in row and pd.notna(row['id']) else f"ROW_{idx}"
        
        if bgm_id in processed_ids:
            print(f"跳过 ID {bgm_id} （已处理）")
            continue

        jp_name = str(row["日文名"]).strip() if pd.notna(row["日文名"]) else ""
        cn_name = str(row["中文名"]).strip() if pd.notna(row["中文名"]) else ""

        if not jp_name and not cn_name:
            print(f"跳过 ID {bgm_id}：日文名和中文名均为空")
            append_unmatched_to_excel(f"ID_{bgm_id}_空名称", unmatched_file)
            continue

        print(f"\n正在匹配 ID {bgm_id} (日文名: '{jp_name}', 中文名: '{cn_name}')")

        best_match = None
        best_score = -1.0  # 初始化最高得分
        match_source = ""

        # 尝试匹配日文名
        if jp_name:
            jp_matches = search_ym_top_matches(jp_name, token_ref)
            if jp_matches and jp_matches[0]["score"] > best_score:
                best_match = jp_matches[0]
                best_score = best_match["score"]
                match_source = "日文名"

        # 尝试匹配中文名
        if cn_name:
            cn_matches = search_ym_top_matches(cn_name, token_ref)
            if cn_matches and cn_matches[0]["score"] > best_score:
                best_match = cn_matches[0]
                best_score = best_match["score"]
                match_source = "中文名"
                
        if best_match:
            row_list: List[Dict[str, Any]] = []
            # ---- 公司信息处理 ----------------------------------------
            org_id = str(best_match.get("orgId", ""))
            org_info = None  # type: Optional[Dict[str, Any]]

            if org_id:
                should_retry = False
                if org_id in processed_orgs:
                    # 信息不完整时重试 (最多 3 次)
                    existing = processed_orgs[org_id]["info"]
                    if not existing.get("website") or not existing.get("description"):
                        should_retry = True
                        processed_orgs[org_id]["retry_count"] += 1
                else:
                    should_retry = True
                    processed_orgs[org_id] = {"info": {}, "retry_count": 1}

                if should_retry and processed_orgs[org_id]["retry_count"] <= 3:
                    org_info = get_organization_details(org_id, token_ref["value"])
                    if org_info:
                        processed_orgs[org_id]["info"] = org_info
                        append_org_to_excel(org_info, org_output_file)
                else:
                    org_info = processed_orgs[org_id]["info"]

            # ---- 组装行数据 -----------------------------------------
            row_data = {
                "bgm_id": bgm_id,
                "bgm游戏": jp_name if jp_name else cn_name, # 使用非空的原始名称作为bgm游戏
                "日文名 (原始)": jp_name,
                "中文名 (原始)": cn_name,
                "name": best_match["name"],
                "chineseName": best_match["chineseName"],
                "ym_id": best_match["ym_id"],
                "score": best_match["score"],
                "orgId": org_id,
                "orgName": (org_info or {}).get("name", best_match.get("orgName", "")),
                "orgWebsite": (org_info or {}).get("website", best_match.get("orgWebsite", "")),
                "orgDescription": (org_info or {}).get("description", best_match.get("orgDescription", "")),
                "匹配来源": match_source
            }
            row_list.append(row_data)
            print(f" - 匹配成功：{best_match['name']} (得分: {best_match['score']})")

            append_to_excel(row_list, output_file)
        else:
            print(" - 未匹配到任何项")
            append_unmatched_to_excel(f"ID_{bgm_id}_未匹配", unmatched_file)

        # 避免触发接口限流
        time.sleep(0.05)

    print("\n所有匹配结果已保存。🎉")

###############################################################################
# 二次匹配：月幕 -> Bangumi 额外信息
###############################################################################

def calculate_similarity(str1: str, str2: str) -> float:
    """利用 ``difflib.SequenceMatcher`` 计算字符串相似度。"""
    return SequenceMatcher(None, str1.lower(), str2.lower()).ratio()

def match_ym_with_bangumi(
    ym_file: str = "ymgames_matched.xlsx",
    bangumi_file: str = "processed_games_test5.xlsx",
    output_file: str = "ym_bangumi_matched.csv"
) -> None:
    """ 
    按名称相似度将 **月幕游戏** 与 **Bangumi 游戏** 对齐，并输出 CSV 文件。
    """
    print("开始匹配月幕游戏与 Bangumi 游戏…")

    # 1. 读取两侧数据
    ym_df = pd.read_excel(ym_file)
    bg_df = pd.read_excel(bangumi_file)

    results = []

    # 2. 遍历月幕条目
    for _, ym_row in ym_df.iterrows():
        ym_name = ym_row["name"]
        ym_cn_name = ym_row["chineseName"]
        ym_id = ym_row["ym_id"]

        best_match, best_score = None, 0.0
        for _, bg_row in bg_df.iterrows():
            score = calculate_similarity(ym_name, bg_row["游戏名称"])
            if score > best_score:
                best_match, best_score = bg_row, score

        if best_match is not None and best_score >= 0.8:
            results.append({
                "ym_id": ym_id,
                "ym_name": ym_name,
                "ym_chinese_name": ym_cn_name,
                "bangumi_id": best_match.get("游戏ID", ""),
                "bangumi_name": best_match["游戏名称"],
                "bangumi_score": best_match.get("评分", ""),
                "bangumi_rank": best_match.get("排名", ""),
                "bangumi_votes": best_match.get("投票数", ""),
                "bangumi_summary": best_match.get("简介", ""),
                "match_score": round(best_score, 4)
            })
            print(f"匹配成功：{ym_name} -> {best_match['游戏名称']} (得分: {best_score:.4f})")

    pd.DataFrame(results).to_csv(output_file, index=False, encoding="utf-8-sig")
    print(f"\n匹配结果已保存到：{output_file}  (共 {len(results)} 条)")

###############################################################################
# 入口
###############################################################################

if __name__ == "__main__":
    # ① Bangumi -> 月幕 首次匹配
    match_bgm_games_and_save()

    # ② 月幕 -> Bangumi 二次匹配
    match_ym_with_bangumi()
